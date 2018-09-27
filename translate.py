# coding=gbk
import urllib
import urllib.request
import json
import time
import hashlib
import openpyxl


class YouDaoFanyi:
    def __init__(self, appKey, appSecret):
        self.url = 'https://openapi.youdao.com/api/'
        self.headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.109 Safari/537.36",
        }
        self.appKey = '****************'  # 应用id
        self.appSecret = '*************'  # 应用密钥
        self.langFrom = 'auto'   # 翻译前文字语言,auto为自动检查
        self.langTo = 'auto'     # 翻译后文字语言,auto为自动检查

    def getUrlEncodedData(self, queryText):
        '''
        将数据url编码
        :param queryText: 待翻译的文字
        :return: 返回url编码过的数据
        '''
        salt = str(int(round(time.time() * 1000)))  # 产生随机数 ,其实固定值也可以,不如"2"
        sign_str = self.appKey + queryText + salt + self.appSecret
    
        sign = hashlib.md5(sign_str.encode("utf-8")).hexdigest()
        payload = {
            'q': queryText,
            'from': self.langFrom,
            'to': self.langTo,
            'appKey': self.appKey,
            'salt': salt,
            'sign': sign
        }

        # 注意是get请求，不是请求
        data = urllib.parse.urlencode(payload)
        return data

    def parseHtml(self, html):
        '''
        解析页面，输出翻译结果
        :param html: 翻译返回的页面内容
        :return: None
        '''
        data = json.loads(html)
       # print ('-' * 10)
        translationResult = data['translation']
       
        if isinstance(translationResult, list):
            translationResult = translationResult[0]
        print (translationResult)
        if "basic" in data:
            youdaoResult = data['basic']
           # print ('有道词典结果')
            return youdaoResult 
        #print ('-' * 10)

    def translate(self, queryText):
        data = self.getUrlEncodedData(queryText)  # 获取url编码过的数据
        target_url = self.url + '?' + data    # 构造目标url
       # print(target_url)
        request = urllib.request.Request(target_url, headers=self.headers)  # 构造请求
        response = urllib.request.urlopen(request)  # 发送请求
        return self.parseHtml(response.read().decode())    # 解析，显示翻译结果


if __name__ == "__main__":

    path="C:/Users/*********************"

    fanyi = YouDaoFanyi(appKey, appSecret)
       
    wb = openpyxl.load_workbook(path)
    sheet = wb.get_sheet_by_name('Sheet1')
    
    i=0
    for row in sheet.rows:
        i+=1
       # print(row[0].value, "\t", end="")	   
        basicJson=fanyi.translate(row[0].value.strip())
        
        if basicJson:
            try:
                sheet.cell(column=2,row=i).value="["+basicJson['us-phonetic']+"]"
            except:
                try:
                    sheet.cell(column=2,row=i).value="["+basicJson['uk-phonetic']+"]"
                except:
                    print("error1")
            try:
                sheet.cell(column=3,row=i).value=basicJson['explains'][0]
            except:
                print("error2")
		       
       # print(basicJson['explains'][0])
       # print(basicJson['us-phonetic'])
    print()
    
    wb.save(path)